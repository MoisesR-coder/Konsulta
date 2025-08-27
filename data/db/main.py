from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from datetime import datetime, timedelta
import tempfile
import os
import shutil
from typing import Optional, List
import jwt
from passlib.context import CryptContext
from pydantic import BaseModel
from sqlalchemy import or_, desc, asc
import logging
import uuid
import time
from sqlalchemy.orm import Session
from database import get_db, create_tables, test_connection, User, ProcessingHistory
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Configuración de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuración de la aplicación
app = FastAPI(
    title="Procesador de Plantillas Excel",
    description="API para procesar archivos Excel de pensiones y generar plantillas de dispersión",
    version="1.0.0"
)

# Configuración CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://31.220.98.150:81", "http://31.220.98.150"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuración de seguridad
SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "30"))

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Modelos Pydantic
class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class ProcessingResult(BaseModel):
    id: str
    filename: str
    original_filename: str
    processed_filename: Optional[str]
    rows_processed: int
    created_at: datetime
    status: str

class PaginatedProcessingResult(BaseModel):
    items: List[ProcessingResult]
    total: int
    page: int
    size: int
    pages: int

# Inicializar base de datos MySQL
def init_db():
    try:
        # Verificar conexión
        if not test_connection():
            logger.error("No se pudo conectar a MySQL. Verifica la configuración.")
            return False
            
        # Crear tablas
        create_tables()
        logger.info("Tablas creadas exitosamente")
        
        # Crear usuario por defecto
        db = next(get_db())
        try:
            existing_user = db.query(User).filter(User.username == "admin").first()
            if not existing_user:
                hashed_password = pwd_context.hash("admin123")
                default_user = User(
                    username="admin",
                    hashed_password=hashed_password
                )
                db.add(default_user)
                db.commit()
                logger.info("Usuario admin creado")
        finally:
            db.close()
            
        return True
    except Exception as e:
        logger.error(f"Error inicializando base de datos: {e}")
        return False

# Funciones de autenticación
def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def authenticate_user(username: str, password: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == username).first()
    
    if not user:
        return False
    if not verify_password(password, user.hashed_password):
        return False
    return {"id": user.id, "username": user.username}

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security), db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise HTTPException(status_code=401, detail="Token inválido")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token inválido")
    
    user = db.query(User).filter(User.username == username).first()
    
    if user is None:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return {"id": user.id, "username": user.username}

# Cargar variables de entorno
load_dotenv()

# Inicializar base de datos al inicio
if not init_db():
    logger.error("Error crítico: No se pudo inicializar la base de datos MySQL")
    exit(1)

@app.get("/")
async def root():
    return {"message": "API de Procesador de Plantillas Excel"}

@app.post("/auth/login", response_model=Token)
async def login(user_data: UserLogin, db: Session = Depends(get_db)):
    user = authenticate_user(user_data.username, user_data.password, db)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciales incorrectas",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

# Funciones de procesamiento de Excel
def find_col(df, keywords):
    """Encuentra la columna que contiene las palabras clave (adaptado del script original)"""
    for kw in keywords:
        for c in df.columns:
            if kw in str(c):
                return c
    return None

def process_excel_file(file_path: str, output_path: str):
    """Procesa el archivo Excel y genera la plantilla de dispersión"""
    try:
        # Leer el archivo Excel con header en la fila 8 (index 7) como en el script original
        try:
            df = pd.read_excel(file_path, sheet_name="ADMON. PENSION", header=7)
            logger.info(f"Archivo leído con header en fila 8, {len(df)} filas")
        except Exception:
            # Si no funciona con sheet específica, intentar con la primera hoja
            try:
                df = pd.read_excel(file_path, header=7)
                logger.info(f"Archivo leído con header en fila 8 (primera hoja), {len(df)} filas")
            except Exception:
                # Último intento: leer normalmente
                df = pd.read_excel(file_path)
                logger.info(f"Archivo leído normalmente, {len(df)} filas")
        
        # Normalizar columnas como en el script original
        df.columns = [str(c).strip().upper().replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U").replace("Ñ","N") for c in df.columns]
        logger.info(f"Columnas normalizadas: {list(df.columns)}")
        
        # Buscar columnas específicas usando la misma lógica del script original
        name_col = find_col(df, ["NOMBRE","NOMBRE COMPLETO","NOMBRECOMPLETO","APELLIDO","EMPLEADO"])
        clabe_col = find_col(df, ["CLABE","CLABEINTERBANCARIA","CLABE INTERBANCARIA","CUENTA","BANCO"])
        amount_col = find_col(df, ["NETO","NETO A DEPOSITAR","MONTO","IMPORTE","PENSION","PAGO","CANTIDAD"])
        
        if not all([name_col, clabe_col, amount_col]):
            missing = []
            if not name_col: missing.append('nombre')
            if not clabe_col: missing.append('CLABE')
            if not amount_col: missing.append('importe')
            logger.error(f"Columnas encontradas - Nombre: {name_col}, CLABE: {clabe_col}, Importe: {amount_col}")
            raise ValueError(f"No se encontraron las columnas: {', '.join(missing)}")
        
        # Crear DataFrame de salida como en el script original
        df_out = pd.DataFrame()
        df_out["Nombre"] = df[name_col] if name_col else ""
        df_out["Clabe"] = df[clabe_col] if clabe_col else ""
        df_out["Monto"] = df[amount_col] if amount_col else ""
        df_out["Concepto"] = "PENSION POR RENTA VITALICIA"
        
        # Limpieza básica como en el script original
        # Convertir CLABEs de notación científica a formato correcto
        df_out["Clabe"] = df_out["Clabe"].apply(lambda x: f"{int(float(x)):018d}" if pd.notna(x) and str(x) != 'nan' else "")
        df_out["Monto"] = pd.to_numeric(df_out["Monto"], errors="coerce").round(2)
        
        # Filtrar filas válidas
        df_out = df_out.dropna(subset=["Nombre","Clabe","Monto"])
        
        # Filtrar filas de totales que no queremos
        totales_keywords = ['NETO A DEPOSITAR', 'COMISION', 'SUBTOTAL', 'IVA', 'TOTAL']
        df_out = df_out[~df_out['Nombre'].astype(str).str.upper().isin([k.upper() for k in totales_keywords])]
        
        df_clean = df_out  # Para mantener compatibilidad con el resto del código
        
        logger.info(f"Datos procesados: {len(df_clean)} filas válidas")
        
        # Crear archivo de salida con formato específico como en el script original
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Definir estilos como en el script original
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        border = Border(
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            left=Side(style='thin'),
            right=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center')
        
        # Escribir encabezados en la primera fila
        headers = ['Nombre', 'Clabe', 'Monto', 'Concepto']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Los datos ahora comienzan desde la fila 2, no necesitamos celda vacía
        
        # Escribir los datos a partir de la fila 2
        for row_idx, (_, row_data) in enumerate(df_clean.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=row_data['Nombre'])
            ws.cell(row=row_idx, column=2, value=row_data['Clabe'])
            
            # Aplicar formato de número con separadores de miles a los montos
            monto_cell = ws.cell(row=row_idx, column=3, value=row_data['Monto'])
            monto_cell.number_format = '#,##0.00'
            
            # Escribir el concepto en la columna 4
            ws.cell(row=row_idx, column=4, value=row_data['Concepto'])
            
            # Aplicar bordes a todas las celdas de datos
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = border
        
        # Ajustar ancho de columnas con espaciado mejorado
        ws.column_dimensions['A'].width = 35  # Nombre
        ws.column_dimensions['B'].width = 25  # Clabe
        ws.column_dimensions['C'].width = 18  # Monto
        ws.column_dimensions['D'].width = 40  # Concepto
        ws.column_dimensions['E'].width = 5   # Columna vacía
        
        # Guardar archivo
        wb.save(output_path)
        logger.info(f"Archivo guardado en: {output_path}")
        
        return len(df_clean)
        
    except Exception as e:
        logger.error(f"Error procesando archivo: {str(e)}")
        raise

@app.post("/upload-process", response_model=ProcessingResult)
async def upload_and_process(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Subir y procesar archivo Excel"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos Excel (.xlsx, .xls)")
    
    processing_id = str(uuid.uuid4())
    
    try:
        # Leer contenido del archivo para obtener el tamaño
        file_content = await file.read()
        file_size = len(file_content)
        
        # Crear directorio temporal
        temp_dir = tempfile.mkdtemp()
        input_path = os.path.join(temp_dir, file.filename)
        output_filename = f"plantilla_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
        output_path = os.path.join(temp_dir, output_filename)
        
        # Guardar archivo subido
        with open(input_path, "wb") as buffer:
            buffer.write(file_content)
        
        # Procesar archivo
        rows_processed = process_excel_file(input_path, output_path)
        
        # Guardar en base de datos MySQL
        processing_record = ProcessingHistory(
            id=processing_id,
            filename=output_filename,
            original_filename=file.filename,
            processed_filename=output_filename,
            rows_processed=rows_processed,
            file_size=file_size,
            user_id=None,  # Sin autenticación
            processing_status="completed"
        )
        db.add(processing_record)
        db.commit()
        
        # Mover archivo procesado a directorio permanente
        os.makedirs('processed_files', exist_ok=True)
        final_output_path = os.path.join('processed_files', f"{processing_id}_{output_filename}")
        shutil.move(output_path, final_output_path)
        
        # Limpiar directorio temporal
        shutil.rmtree(temp_dir)
        
        return ProcessingResult(
            id=processing_id,
            filename=output_filename,
            original_filename=file.filename,
            processed_filename=output_filename,
            rows_processed=rows_processed,
            created_at=datetime.now(),
            status="completed"
        )
        
    except Exception as e:
        # Limpiar en caso de error
        if 'temp_dir' in locals():
            shutil.rmtree(temp_dir, ignore_errors=True)
        
        logger.error(f"Error procesando archivo: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error procesando archivo: {str(e)}")

@app.get("/download/{processing_id}")
async def download_file(
    processing_id: str,
    db: Session = Depends(get_db)
):
    """Descargar archivo procesado"""
    processing_record = db.query(ProcessingHistory).filter(
        ProcessingHistory.id == processing_id
    ).first()
    
    if not processing_record:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    file_path = os.path.join('processed_files', f"{processing_id}_{processing_record.filename}")
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Archivo no encontrado en el sistema")
    
    return FileResponse(
        path=file_path,
        filename=processing_record.filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.get("/history", response_model=List[ProcessingResult])
async def get_processing_history(
    db: Session = Depends(get_db)
):
    """Obtener historial de procesamiento"""
    processing_records = db.query(ProcessingHistory).order_by(ProcessingHistory.created_at.desc()).all()
    
    return [
        ProcessingResult(
            id=record.id,
            filename=record.filename,
            original_filename=record.original_filename,
            processed_filename=record.processed_filename,
            rows_processed=record.rows_processed,
            created_at=record.created_at,
            status=record.processing_status
        )
        for record in processing_records
    ]

@app.get("/history/paginated", response_model=PaginatedProcessingResult)
async def get_paginated_processing_history(
    page: int = 1,
    size: int = 10,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    db: Session = Depends(get_db)
):
    """Obtener historial de procesamiento con paginación, filtrado y ordenamiento"""
    
    # Validar parámetros
    if page < 1:
        page = 1
    if size < 1 or size > 100:
        size = 10
    if sort_by not in ["created_at", "filename", "original_filename", "processed_filename", "rows_processed", "processing_status"]:
        sort_by = "created_at"
    if sort_order not in ["asc", "desc"]:
        sort_order = "desc"
    
    # Construir query base
    query = db.query(ProcessingHistory)
    
    # Aplicar filtro de búsqueda si se proporciona
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            or_(
                ProcessingHistory.filename.like(search_filter),
                ProcessingHistory.processing_status.like(search_filter)
            )
        )
    
    # Aplicar ordenamiento
    sort_column = getattr(ProcessingHistory, sort_by)
    if sort_order == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(asc(sort_column))
    
    # Obtener total de registros
    total = query.count()
    
    # Aplicar paginación
    offset = (page - 1) * size
    processing_records = query.offset(offset).limit(size).all()
    
    # Calcular número total de páginas
    pages = (total + size - 1) // size
    
    # Convertir a modelo de respuesta
    items = [
        ProcessingResult(
            id=record.id,
            filename=record.filename,
            original_filename=record.original_filename,
            processed_filename=record.processed_filename,
            rows_processed=record.rows_processed,
            created_at=record.created_at,
            status=record.processing_status
        )
        for record in processing_records
    ]
    
    return PaginatedProcessingResult(
        items=items,
        total=total,
        page=page,
        size=size,
        pages=pages
    )

@app.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now()}