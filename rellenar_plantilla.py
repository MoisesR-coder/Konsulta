#!/usr/bin/env python3
"""Script: rellenar_plantilla.py
Lee el archivo ANEXO A4 (hoja "ADMON. PENSION") y crea un Excel con las columnas:
Nombre, Clabe, Monto, Concepto
Adaptable a cambios menores en nombres de columnas.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from pathlib import Path
from datetime import datetime

file_datos = Path("ANEXO A4 ESFLO MKTING Q1 Agosto 2025.xlsx")
file_plantilla = Path("plantilla-dispersion-actual.xlsx")
out = Path(f"plantilla_{datetime.now().strftime('%Y-%m-%d')}.xlsx")

# Lee con header en la fila 8 (index 7)
df = pd.read_excel(file_datos, sheet_name="ADMON. PENSION", header=7)
# normalizar columnas
df.columns = [str(c).strip().upper().replace("Á","A").replace("É","E").replace("Í","I").replace("Ó","O").replace("Ú","U").replace("Ñ","N") for c in df.columns]

# helper de búsqueda por palabra clave
def find_col(df, keywords):
    for kw in keywords:
        for c in df.columns:
            if kw in c:
                return c
    return None

col_nombre = find_col(df, ["NOMBRE","NOMBRE COMPLETO","NOMBRECOMPLETO","APELLIDO","EMPLEADO"])
col_clabe = find_col(df, ["CLABE","CLABEINTERBANCARIA","CLABE INTERBANCARIA","CUENTA","BANCO"])
col_monto = find_col(df, ["NETO","NETO A DEPOSITAR","MONTO","IMPORTE","PENSION","PAGO","CANTIDAD"])

df_out = pd.DataFrame()
df_out["Nombre"] = df[col_nombre] if col_nombre else ""
df_out["Clabe"] = df[col_clabe] if col_clabe else ""
df_out["Monto"] = df[col_monto] if col_monto else ""
df_out["Concepto"] = "PENSION POR RENTA VITALICIA"

# limpieza basica
# Convertir CLABEs de notación científica a formato correcto
df_out["Clabe"] = df_out["Clabe"].apply(lambda x: f"{int(float(x)):018d}" if pd.notna(x) and str(x) != 'nan' else "")
df_out["Monto"] = pd.to_numeric(df_out["Monto"], errors="coerce").round(2)

df_out = df_out.dropna(subset=["Nombre","Clabe","Monto"])

# Filtrar filas de totales que no queremos
totales_keywords = ['NETO A DEPOSITAR', 'COMISION', 'SUBTOTAL', 'IVA', 'TOTAL']
df_out = df_out[~df_out['Nombre'].astype(str).str.upper().isin([k.upper() for k in totales_keywords])]

# Crear un nuevo workbook con formato similar a la plantilla
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Definir estilos
header_font = Font(bold=True)
header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
border = Border(
    top=Side(style='thin'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)
center_alignment = Alignment(horizontal='center')

# Escribir encabezados en la primera fila
headers = ['Nombre', 'Clabe', 'Monto', 'Concepto']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center_alignment

# Los datos ahora comienzan desde la fila 2, no necesitamos celda vacía

# Escribir los datos a partir de la fila 2
for row_idx, (_, row_data) in enumerate(df_out.iterrows(), start=2):
    ws.cell(row=row_idx, column=1, value=row_data['Nombre'])
    ws.cell(row=row_idx, column=2, value=row_data['Clabe'])
    
    # Aplicar formato de número con separadores de miles a los montos
    monto_cell = ws.cell(row=row_idx, column=3, value=row_data['Monto'])
    monto_cell.number_format = '#,##0.00'
    
    # Escribir el concepto en la columna 4
    ws.cell(row=row_idx, column=4, value=row_data['Concepto'])
    
    # Aplicar bordes a todas las celdas de datos
    for col in range(1, 5):
        ws.cell(row=row_idx, column=col).border = border

# Ajustar ancho de columnas con espaciado mejorado
ws.column_dimensions['A'].width = 35  # Nombre - más espacio para nombres largos
ws.column_dimensions['B'].width = 25  # Clabe - espacio adicional para números largos
ws.column_dimensions['C'].width = 18  # Monto - espacio para cantidades con decimales
ws.column_dimensions['D'].width = 40  # Concepto - más espacio para texto completo
ws.column_dimensions['E'].width = 5   # Columna vacía para separación visual

# Guardar el archivo
wb.save(out)
print(f"Archivo generado: {out} con {len(df_out)} filas de datos")
print("Formato aplicado: encabezados en negrita, centrados, con bordes y tabla responsiva")
